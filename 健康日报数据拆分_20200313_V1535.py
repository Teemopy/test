import pandas as pd
import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
import win32com.client

passkey = '0525'
r_csv = r'\20200525北京链家健康日报填报情况明细_20200525_v1222.csv'
path = r'E:\2020健康日报数据拆分'+'\\'+passkey
path1 = r'E:\2020健康日报数据拆分'+'\\'+passkey+' - 副本'

Daytime = '_2020'+passkey+'_V'
# 读取数据
data0 = pd.read_csv(path + r_csv, engine='python', encoding='utf-8-sig')
data = data0[data0['三级组织'] == '职能']
data_y = data0[data0['三级组织'] == '运营']

gg = data[(data['四级组织'] == '公共关系中心') | (data['员工姓名'] == '刘海琦')]
zc = data[data['四级组织'] == '运营支持中心']
rl = data[data['四级组织'] == '人力资源中心']
cw = data[data['四级组织'] == '财务中心']
sc = data[data['四级组织'] == '市场中心']
yy = data[data['四级组织'] == '运营管理中心']
hg = data[data['四级组织'] == '合规监察及道德廉洁建设部']
kf = data[data['四级组织'] == '客户服务中心']
sl = data_y[data_y['四级组织'] == '社区链接中心']
sb = data_y[data_y['四级组织'] == '商办平台中心']
ac = data[data['四级组织'] == '服务品质中心']
cp = data[data['四级组织'] == '产品赋能中心']
ds = data[data['四级组织'] == '店面设计工程中心']
dg = data[data['四级组织'] == '店面管理中心']
fw = data[data['四级组织'] == '法务中心']


curr_time = datetime.datetime.now()
h = str(curr_time.hour)
m = str(curr_time.minute)

if len(m) == 1:
    m = '0' + m
else:
    m = m
if len(h) == 1:
    h = '0' + h
else:
    h = h

zcpath = path + r'\运营支持中心健康日报明细' + Daytime + h + m + '.xlsx'
cwpath = path1 + r'\财务中心健康日报明细' + Daytime + h + m + '.xlsx'
rlpath = path1 + r'\人力资源中心健康日报明细' + Daytime + h + m + '.xlsx'
scpath = path1 + r'\市场中心健康日报明细' + Daytime + h + m + '.xlsx'
yypath = path1 + r'\运营管理中心健康日报明细' + Daytime + h + m + '.xlsx'
ggpath = path + r'\公共关系中心健康日报明细' + Daytime + h + m + '.xlsx'
hgpath = path + r'\合规监察及道德廉洁建设部健康日报明细' + Daytime + h + m + '.xlsx'
kfpath = path + r'\客户服务中心健康日报明细' + Daytime + h + m + '.xlsx'
slpath = path + r'\社区链接中心健康日报明细' + Daytime + h + m + '.xlsx'
sbpath = path + r'\商办平台中心健康日报明细' + Daytime + h + m + '.xlsx'
acpath = path + r'\服务品质中心健康日报明细' + Daytime + h + m + '.xlsx'
cppath = path + r'\产品赋能中心健康日报明细' + Daytime + h + m + '.xlsx'
dspath = path + r'\店面设计工程中心健康日报明细' + Daytime + h + m + '.xlsx'
dgpath = path + r'\店面管理中心健康日报明细' + Daytime + h + m + '.xlsx'
fwpath = path + r'\法务中心健康日报明细' + Daytime + h + m + '.xlsx'

lst = [cwpath, rlpath, scpath, yypath, ggpath, hgpath, kfpath, slpath, acpath, cppath,
       dspath, dgpath, fwpath, zcpath, sbpath]

zc.to_excel(zcpath, index=False)
cw.to_excel(cwpath, index=False)
rl.to_excel(rlpath, index=False)
sc.to_excel(scpath, index=False)
yy.to_excel(yypath, index=False)
gg.to_excel(ggpath, index=False)
hg.to_excel(hgpath, index=False)
kf.to_excel(kfpath, index=False)
sl.to_excel(slpath, index=False)
sb.to_excel(sbpath, index=False)
ac.to_excel(acpath, index=False)
cp.to_excel(cppath, index=False)
ds.to_excel(dspath, index=False)
dg.to_excel(dgpath, index=False)
fw.to_excel(fwpath, index=False)

xcl = win32com.client.Dispatch("Excel.Application")


def exc(filename):
    # 定义修改格式函数
    def gs(a, b):
        font = Font(name='等线', size=11, bold=None)
        border_set = Border(left=Side(style=None),
                            right=Side(style=None),
                            top=Side(style=None),
                            bottom=Side(style=None))

        ws.cell(row=a, column=b).font = font
        ws.cell(row=a, column=b).border = border_set
        return

    # 读取文件
    wb = load_workbook(filename)
    ws = wb['Sheet1']

    rows = ws.max_row
    cols = ws.max_column

    for i in range(1, rows+1):
        for j in range(1, cols+1):
            gs(i, j)

    wb.save(filename)
    wb = xcl.Workbooks.Open(filename, False, False, None, '')
    xcl.DisplayAlerts = False
    wb.SaveAs(filename, None, passkey)

    print(datetime.datetime.now(), filename)
    return


# 执行excel格式调整部分
for excel_path in lst:
    exc(excel_path)


xcl.Quit()
